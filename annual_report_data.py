
from openpyxl import load_workbook
from bcompiler.utils import project_data_from_master
import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, IconSet, FormatObject


'''function for calculating if confidence has increased decreased'''
def up_or_down(latest_dca, last_dca):

    if latest_dca == last_dca:
        return (int(0))
    elif latest_dca != last_dca:
        if last_dca == 'Green':
            if latest_dca != 'Amber/Green':
                return (int(-1))
        elif last_dca == 'Amber/Green':
            if latest_dca == 'Green':
                return (int(1))
            else:
                return (int(-1))
        elif last_dca == 'Amber':
            if latest_dca == 'Green':
                return (int(1))
            elif latest_dca == 'Amber/Green':
                return (int(1))
            else:
                return (int(-1))
        elif last_dca == 'Amber/Red':
            if latest_dca == 'Red':
                return (int(-1))
            else:
                return (int(1))
        else:
            return (int(1))


def cal_date_difference(milestone_date, old_milestone_date):
    try:
        time_delta = (milestone_date - old_milestone_date).days
    except TypeError:
        time_delta = 0
    return time_delta

'''function that places all information into the summary dashboard sheet'''
def placing_excel(dict_one, dict_two):

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=2).value
        print(project_name)
        if project_name in dict_one:
            dca_one = dict_one[project_name]['DCA']
            try:
                dca_two = dict_two[project_name]['DCA']
                change = up_or_down(dca_one, dca_two)
                ws.cell(row=row_num, column=4).value = change
            except KeyError:
                ws.cell(row=row_num, column=4).value = 'NEW'
            ws.cell(row=row_num, column=5).value = dict_one[project_name]['DCA']

            start_date_one = dict_one[project_name]['Start Date']
            ws.cell(row=row_num, column=6).value = start_date_one
            try:
                start_date_two = dict_two[project_name]['Start Date']
                s_date_diff = cal_date_difference(start_date_one, start_date_two)
                ws.cell(row=row_num, column=7).value = s_date_diff
            except KeyError:
                ws.cell(row=row_num, column=7).value = 0

            end_date_one = dict_one[project_name]['End Date']
            ws.cell(row=row_num, column=8).value = end_date_one
            try:
                end_date_two = dict_two[project_name]['End Date']
                e_date_diff = cal_date_difference(end_date_one, end_date_two)
                ws.cell(row=row_num, column=9).value = e_date_diff
            except KeyError:
                ws.cell(row=row_num, column=9).value = 0

            ws.cell(row=row_num, column=10).value = dict_one[project_name]['18/19 baseline']
            ws.cell(row=row_num, column=11).value = dict_one[project_name]['18/19 forecast']
            ws.cell(row=row_num, column=12).value = dict_one[project_name]['18/19 variance']
            wlc_one = dict_one[project_name]['WLC baseline']
            ws.cell(row=row_num, column=13).value = wlc_one
            try:
                wlc_two = dict_two[project_name]['WLC baseline']
                wlc_diff = wlc_one - wlc_two
                ws.cell(row=row_num, column=14).value = wlc_diff
            except KeyError:
                ws.cell(row=row_num, column=14).value = 0


    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=2).value
        if project_name in dict_two:
            ws.cell(row=row_num, column=3).value = dict_two[project_name]['DCA']

    # Highlight cells that contain RAG text, with background and text the same colour. column E.
    ag_text = Font(color="00a5b700")
    ag_fill = PatternFill(bgColor="00a5b700")
    dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Green",e1)))']
    ws.conditional_formatting.add('e1:e100', rule)

    ar_text = Font(color="00f97b31")
    ar_fill = PatternFill(bgColor="00f97b31")
    dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Red",e1)))']
    ws.conditional_formatting.add('e1:e100', rule)

    red_text = Font(color="00fc2525")
    red_fill = PatternFill(bgColor="00fc2525")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Red",E1)))']
    ws.conditional_formatting.add('E1:E100', rule)

    green_text = Font(color="0017960c")
    green_fill = PatternFill(bgColor="0017960c")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Green",e1)))']
    ws.conditional_formatting.add('E1:E100', rule)

    amber_text = Font(color="00fce553")
    amber_fill = PatternFill(bgColor="00fce553")
    dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber",e1)))']
    ws.conditional_formatting.add('e1:e100', rule)

    # Highlight cells that contain RAG text, with background and black text columns G to L.
    ag_text = Font(color="000000")
    ag_fill = PatternFill(bgColor="00a5b700")
    dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Green",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    ar_text = Font(color="000000")
    ar_fill = PatternFill(bgColor="00f97b31")
    dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Red",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    red_text = Font(color="000000")
    red_fill = PatternFill(bgColor="00fc2525")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Red",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    green_text = Font(color="000000")
    green_fill = PatternFill(bgColor="0017960c")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Green",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    amber_text = Font(color="000000")
    amber_fill = PatternFill(bgColor="00fce553")
    dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    # highlighting new projects
    red_text = Font(color="00fc2525")
    white_fill = PatternFill(bgColor="000000")
    dxf = DifferentialStyle(font=red_text, fill=white_fill)
    rule = Rule(type="containsText", operator="containsText", text="NEW", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("NEW",F1)))']
    ws.conditional_formatting.add('D1:D100', rule)

    # assign the icon set to a rule
    first = FormatObject(type='num', val=-1)
    second = FormatObject(type='num', val=0)
    third = FormatObject(type='num', val=1)
    iconset = IconSet(iconSet='3Arrows', cfvo=[first, second, third], percent=None, reverse=None)
    rule = Rule(type='iconSet', iconSet=iconset)
    ws.conditional_formatting.add('D1:D100', rule)

    return wb


'''1) Provide file path to empty dashboard document'''
wb = load_workbook(
    'C:\\Users\\Standalone\\Will\\ipa_annual_report_dashboard_master.xlsx')
ws = wb.active

'''2) Provide file path to master data sets'''
data_one = project_data_from_master(
    'C:\\Users\\Standalone\\Will\\DfT AR 2019 Data.xlsx')
data_two = project_data_from_master(
    'C:\\Users\\Standalone\\Will\\ipa_annual_report_2018.xlsx')

p_names = list(data_one.keys())

wb = placing_excel(data_one, data_two)

'''4) provide file path and specific name of output file.'''
wb.save(
    'C:\\Users\\Standalone\\Will\\test.xlsx')